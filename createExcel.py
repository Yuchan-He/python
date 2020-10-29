# -*- coding: utf-8 -*-
"""
create excel
"""

from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 300
ws.cell(row=1, column=2, value='abc')
wb.save('test2.xlsx')

"""
open excel
"""
wb = load_workbook('test1.xlsx')
ws = wb.active
ws['A2'] = 'write in'
wb.save('test1.xlsx')

# 二次元配列を巡回
squares = [i**2 for i in range(5)]
print(squares)

squares = []
for i in range(5):
    squares.append(i**2)

print(squares)
