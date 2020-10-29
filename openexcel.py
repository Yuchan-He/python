from openpyxl import load_workbook

wb = load_workbook('tes1.xlsx')
# 作成したexcelのシートを選択する
ws = wb.active
ws['A2'] = 'write in'
wb.save('test1.xlsx')
