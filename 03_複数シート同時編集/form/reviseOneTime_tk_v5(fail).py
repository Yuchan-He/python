# -*- coding: utf-8 -*-
"""
複数シートを同時入力
"""
import glob
from openpyxl import load_workbook
import tkinter as tk
import tkinter.messagebox as tkm

def batExcel(term, master):
	# 記入するマスタのシートを指定する	
	Sheet = master

	# 複数のファイルを選択
	files = glob.glob('*.xlsx')
	successFile = []
	failFile = []

	# すべてのファイルをループする
	for file in files:
		wb = load_workbook(file)
		sheetnames = (wb.sheetnames)
		
		# マスタSheet表が存在するかどうか判断
		if Sheet in sheetnames:
			# sheetを選択する
			ws = wb[Sheet]
			# cellを選択する
			ws.cell(row=1 , column=2 , value=Entry1.get())
			wb.save(file)
			wb.close()
			successFile.append(file)
		else:
			failFile.append(file)

	# 入力成功、失敗のEXCELの名前をreturn
	return successFile,failFile


# ボダンをクリックするとき、発生する事件
def button_click(batExcel(text, master)):
	# ボタンを押すたびに、リストの中の元データをクリアする
	ListBox1.delete(0, tk.END)
	ListBox2.delete(0, tk.END)
	files = batExcel(text, master)
	for i in successFile:
		ListBox1.insert(tk.END, i)
	for j in failFile:
		ListBox2.insert(tk.END, j)

# GUI画面を作成
root = tk.Tk()

# タイトルを設定
root.title("会計期間")

# 画面のサイズを設定
root.geometry('400x500')

# 会計期間labelを設定
Static1 = tk.Label(text='YYYYMM形式で、会計期間をご入力ください')
Static1.pack()

# 会計期間入力欄を設定
Entry1 = tk.Entry()
Entry1.pack()

# マスタシートlabelを設定
Static2 = tk.Label(text='更新するマスタ名前をご入力ください')
Static2.pack()

# マスタシート入力欄を設定
Entry2 = tk.Entry()
Entry2.pack()

# 実行ボタンを設置
Button = tk.Button(text='実行ボタン', command=lambda: button_click(Entry1.get(), Entry2.get()))
Button.pack()

# 導入成功リストボックスを設置
ListBox1 = tk.Listbox(width=50, height=10)
ListBox1.insert(tk.END,'会計期間更新成功EXCEL:')
ListBox1.pack()

# 導入失敗リストボックスを設置
ListBox2 = tk.Listbox(width=50, height=10)
ListBox2.insert(tk.END,'マスタSheetが存在しないため、会計期間更新失敗EXCEL:')
ListBox2.pack()

root.mainloop()



