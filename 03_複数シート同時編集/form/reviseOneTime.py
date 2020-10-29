# -*- coding: utf-8 -*-
"""
複数シートを同時編集
"""
import glob
from openpyxl import load_workbook

# 会計期間を入力する
# term = input('会計期間:')
term = '123'
Sheet = 'Sheet'

# 複数のファイルを選択
files = glob.glob('*.xlsx')
print(files)

for file in files:
	# bookを選択する
	wb = load_workbook(file)
	print(file)
	sheetnames = (wb.sheetnames)
	if Sheet in sheetnames:
		# sheetを選択する
		ws = wb['Sheet']

		# cellを選択する
		ws.cell(row=1 , column=2 , value=term)
		wb.save(file)
		wb.close()
	else:
		print('Sheet 表が持っていない')

