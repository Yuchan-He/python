# -*- coding: utf-8 -*-
"""
複数シートを同時編集
"""
import glob
from openpyxl import load_workbook
import tkinter as tk
import tkinter.messagebox as tkm

def batExcel(term):
	# 会計期間を入力する
	# term = input('会計期間:')
	# term = '123'
	Sheet = 'Sheet'

	# 複数のファイルを選択
	files = glob.glob('*.xlsx')
	print(files)

	for file in files:
		# bookを選択する
		wb = load_workbook(file)
		print(file)
		sheetnames = (wb.sheetnames)
		if Sheet in sheetnames:
			# sheetを選択する
			ws = wb['Sheet']

			# cellを選択する
			ws.cell(row=1 , column=2 , value=Entry1.get())
			wb.save(file)
			wb.close()
		else:
			print('Sheet 表が持っていない')


# define the button
def button_click(text):
	# tkm.showinfo('会計期間更新しました:', text, 'hi')
	tkm.showinfo('会計期間更新しました:', text)
	batExcel(text)

# create a window
root = tk.Tk()

# set a title
root.title("会計期間")

# set size of the window
root.geometry('400x300')

# create a label
Static1 = tk.Label(text='会計期間をご入力ください')
Static1.pack()

# create a entry
Entry1 = tk.Entry()
Entry1.pack()

# get the value in the Entry1
Entry1_value = Entry1.get()

# set the button
# lambda只能有一个参数
Button = tk.Button(text='実行ボタン', command=lambda: button_click(Entry1.get()))
Button.pack()

# active the window
root.mainloop()



