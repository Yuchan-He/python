# -*- coding: utf-8 -*-
"""
複数シートを同時入力
"""
import glob
from openpyxl import load_workbook
import tkinter as tk
import tkinter.messagebox as tkm

def batExcel(term):
	# 記入するマスタのシートを指定する	
	Sheet = 'Sheet'

	# 複数のファイルを選択
	files = glob.glob('*.xlsx')
	successFile = []
	failFile = []

	# すべてのファイルをループする
	for file in files:
		wb = load_workbook(file)
		sheetnames = (wb.sheetnames)
		
		# マスタSheet表が存在するかどうか判断
		if Sheet in sheetnames:
			# sheetを選択する
			ws = wb['Sheet']
			# cellを選択する
			ws.cell(row=1 , column=2 , value=Entry1.get())
			wb.save(file)
			wb.close()
			successFile.append(file)
		else:
			failFile.append(file)

	# マスタ表が存在しない表を出力する
	# print(f"{failFile}にはマスタ表が存在しておりません")
	return successFile,failFile



# ボダンをクリックするとき、発生する事件
def button_click(text):
	files = batExcel(text)
	for i in range(len(files[0])):
		ListBox1.insert(tk.END, files[0][i])
	for j in files[1]:
		ListBox2.insert(tk.END, j)
	# tkm.showinfo('会計期間', files)

# GUI画面を作成
root = tk.Tk()

# タイトルを設定
root.title("会計期間")

# 画面のサイズを設定
root.geometry('400x300')

# labelを設定
Static1 = tk.Label(text='YYYYMM形式で、会計期間をご入力ください')
Static1.pack()

# 入力欄を設定
Entry1 = tk.Entry()
Entry1.pack()

# 入力したデータを取得
Entry1_value = Entry1.get()

# 実行ボタンを設置
Button = tk.Button(text='実行ボタン', command=lambda: button_click(Entry1.get()))
Button.pack()

# 導入成功リストボックスを設置
ListBox1 = tk.Listbox(width=50, height=5)
ListBox1.insert(tk.END,'会計期間更新成功EXCEL:')
ListBox1.pack()

# 導入失敗リストボックスを設置
ListBox2 = tk.Listbox(width=50, height=5)
ListBox2.insert(tk.END,'マスタSheetが存在しないため、会計期間更新失敗EXCEL:')
ListBox2.pack()

root.mainloop()



