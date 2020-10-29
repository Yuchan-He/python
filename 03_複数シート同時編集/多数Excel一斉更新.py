# -*- coding: utf-8 -*-
"""
複数シートを同時入力
"""
import glob
from openpyxl import load_workbook
import tkinter as tk
import tkinter.messagebox as tkm

def batExcel(term, sheet, cell):
	# 記入するマスタのシートを指定する	
	Sheet = sheet

	# 複数のファイルを選択
	files = glob.glob('*.xlsx')
	successFile = []
	failFile = []

	# すべてのファイルをループする
	for file in files:
		wb = load_workbook(file)
		# すべてのシートを読み取る
		sheetnames = (wb.sheetnames)
		
		# マスタSheet表が存在するかどうか判断
		if Sheet in sheetnames:
			# sheetを選択する
			ws = wb[Sheet]
			# cellを選択する
			ws[cell] = term
			wb.save(file)
			wb.close()
			successFile.append(file)
		else:
			failFile.append(file)

	# 入力成功、失敗のEXCELの名前をreturn
	return successFile,failFile


# ボダンをクリックするとき、発生する事件
# 引数三つ、text -> 入力したいデータ、master ->　入力したいシート、cell ->入力したいセル
def button_click(text, master, cell):
	# ボタンを押すたびに、リストの中の元データをクリアする
	ListBox1.delete(1, tk.END)
	ListBox2.delete(1, tk.END)
	files = batExcel(text, master, cell)
	for i in range(len(files[0])):
		ListBox1.insert(tk.END, files[0][i])
	for j in files[1]:
		ListBox2.insert(tk.END, j)

# GUI画面を作成
root = tk.Tk()

# タイトルを設定
root.title("会計期間")

# 画面のサイズを設定
root.geometry('400x500')

# 会計期間labelを設定
Static1 = tk.Label(text='YYYYMM形式で、会計期間をご入力ください')
Static1.pack()

# 会計期間入力欄を設定
Entry1_term = tk.Entry()
Entry1_term.pack()

# マスタシートlabelを設定
Static2 = tk.Label(text='更新するマスタ名前をご入力ください')
Static2.pack()

# マスタシート入力欄を設定
Entry2_sheet = tk.Entry()
Entry2_sheet.pack()

# マスタシートのセル場所labelを設定
Static3 = tk.Label(text='更新するセル場所をご入力ください')
Static3.pack()

# マスタシートのセル場所入力欄を設定
Entry3_cell = tk.Entry()
Entry3_cell.pack()

# 実行ボタンを設置
Button = tk.Button(text='実行ボタン', command=lambda: button_click(Entry1_term.get(), Entry2_sheet.get(), Entry3_cell.get()))
Button.pack()

# 導入成功リストボックスを設置
ListBox1 = tk.Listbox(width=50, height=10)
ListBox1.insert(tk.END,'会計期間更新成功EXCEL:')
ListBox1.pack()

# 導入失敗リストボックスを設置
ListBox2 = tk.Listbox(width=50, height=10)
ListBox2.insert(tk.END,'マスタSheetが存在しないため、会計期間更新失敗EXCEL:')
ListBox2.pack()

# GUI画面をActive、例えば、マウントクリック、画面閉じるなどの動作に反応する
root.mainloop()
