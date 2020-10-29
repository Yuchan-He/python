# -*- coding: utf-8 -*-
"""
excel データを転記する
"""
# from openpyxl import load_workbook
import openpyxl
import datetime

# ファイルの場所を特定
filepath = 'C:/xampp/htdocs/study/python/01_データ転記/customers_form.xlsx'
# filepath = 'C:\\xampp\\htdocs\\study\\python\\01_データ転記\\customers_form.xlsx'

# ファイルを読み込む
wb = openpyxl.load_workbook(filename=filepath)

# ファイルシートを取得
sheetnames = wb.sheetnames
print(sheetnames)