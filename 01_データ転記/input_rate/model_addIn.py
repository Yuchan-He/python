# -*- coding: utf-8 -*-
"""
excel データを転記する
"""

#ステップ1｜ライブラリの設定
from openpyxl import load_workbook
import datetime

#ステップ2｜所定フォルダ内の「Book1.xlsm」を指定して読み込む
filepath = 'C:/xampp/htdocs/study/python/01_データ転記/customers_form.xlsx'

wb = load_workbook(filename=filepath)
ws1 = wb['データ']
ws2 = wb['集計']

#ステップ3｜集計範囲の取得
startdate=datetime.datetime(int(ws2['B2'].value) , int(ws2['C2'].value) , int(ws2['D2'].value))
enddate=datetime.datetime(int(ws2['B3'].value) , int(ws2['C3'].value) , int(ws2['D3'].value))

#ステップ4｜エクセルの最終行や最終列の取得
lastrow1=ws1.max_row
lastrow2=ws2.max_row
lastcol2=ws2.max_column

#ステップ5｜「データ」シートを読み込み、2次元配列values1として取得
values1=[[cell.value for cell in row1] for row1 in ws1]

#ステップ6｜条件に合う情報をFor文とIF文で場合分けして取得
for i in range(7, lastrow2+1):
    for j in range(2, lastcol2+1):
        counter = 0
        for k in range(1, lastrow1):
            if  values1[k][1] == ws2.cell(row=i, column=1).value:
                if values1[k][2] == ws2.cell(row=6, column=j).value:
                    torihikidate = values1[k][3]
                    if startdate <= torihikidate <= enddate:
                        kingaku=values1[k][4]
                        counter = counter + int(kingaku)

        #ステップ7｜ステップ6に合わない場合に0を入れる
        if counter is None:
                    counter=0

        #ステップ8｜「集計」シートに結果を書き出す
        ws2.cell(row=i, column=j).value = counter

#ステップ9｜「Book2.xlsm」として所定のフォルダに保存する
newfilepath =  'C:/xampp/htdocs/study/python/01_データ転記/model_addIn.xlsx'
wb.save(newfilepath)
