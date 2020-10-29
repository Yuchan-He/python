# -*- coding: utf-8 -*-
"""
excel データを転記する
"""
# from openpyxl import load_workbook
import openpyxl
import datetime
import io
import sys

# 文字化け解消
sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')

# ファイルの場所を特定
filepath = 'C:/xampp/htdocs/study/python/01_データ転記/customers_form.xlsx'
# filepath = 'C:\\xampp\\htdocs\\study\\python\\01_データ転記\\customers_form.xlsx'

# ファイルを読み込む
wb = openpyxl.load_workbook(filename=filepath)

# ファイルシートを取得
sheetnames = wb.sheetnames
print(sheetnames)
ws1 = wb['データ']
ws2 = wb['集計']
# print("おはよう",encode('Shift_JIS'))
# print sys.getdefaultencoding()
# print(sys.getdefaultencoding())
# mycode = sys.getdefaultencoding()
# print(mycode)
# print('おはよう')
# print('你好')

# 集計範囲条件の取得する。
startdate = datetime.datetime(int(ws2['B2'].value) , int(ws2['C2'].value) , int(ws2['D2'].value))
enddate = datetime.datetime(int(ws2['B3'].value) , int(ws2['C3'].value) , int(ws2['D3'].value))
print(startdate)


# value2 = int(ws2['D7'].value)
# print(value2)

# シートの最終行、最終列を取得
lastRow1 = ws1.max_row
print(lastRow1)
lastRow2 = ws2.max_row
print(lastRow2)
lastCol2 = ws2.max_column
print(lastCol2)


# シートの中値を取得する
# value1 = ws1['S50'].value
# print(value1)

# シートを読み込み、2次元入れる
# 1.读取每一行：for row1 in ws1
# 2.读取每一行中单元格的值：cell.value for cell in row1
# 3. 2)中读取到的数值，保存在[]中
values1 = [[cell.value for cell in row1] for row1 in ws1]
print(values1)

# 条件に合う情報をif分で場合分けして取得
for i in range(7 , lastRow2+1):
	for j in range(2 , lastCol2+1):
		counter = 0
		for k in range(1,lastRow1):
			if values1[k][1] == ws2.cell(row=i , column=1).value:
				if values1[k][2] == ws2.cell(row=6 , column=j).value:
					torihikidate = values1[k][3]
					if startdate <= torihikidate <= enddate:
						kingaku = values1[k][4]
						counter = counter + int(kingaku)
		# if counter is None:
		# 	counter = 0

		ws2.cell(row = i , column = j).value = counter

print(counter)



newfile = 'C:/xampp/htdocs/study/python/01_データ転記/addIn.xlsx'
wb.save(newfile)

# newfile = 'C:/xampp/htdocs/study/python/01_データ転記/customers_form.xlsx'
# wb.save(newfile)




# values2 = [cell.value for cell in row1]
# print(values2)
# for row1 in ws1:
	# pirnt row1

# 等価for文 cell.value for cell in row1
# squares = []
# for cell in row1:
# 	squares.append(cell.value)
# print(squares)
# for cell in row1:
	# print(cell.value)




# value4 = [[i, j] for i in range(10) for j in range(10)]
# print(value4)

# value4 = [[i, j] for i in range(10) for j in i]
# print(value4)

wbNew = openpyxl.Workbook()
wbNew.save('newExcel.xlsx')

