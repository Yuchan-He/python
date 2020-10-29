import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheets = wb.sheetnames
print(sheets)

"""
特定なシートを取得する
"""

sheetOne = wb['data']
print(sheetOne['B1'].value)

"""
row,column を指定して、データを取得
"""

print(sheetOne.cell(row=2, column=3).value)
