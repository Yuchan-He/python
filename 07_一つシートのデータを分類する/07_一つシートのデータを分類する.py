# -*- coding: utf-8 -*-
"""
このモジュールの説明
02_シートコピー.xlsx
"""

import sys 
from openpyxl import load_workbook
import time

# ドラッグでexcelを開く
fp = sys.argv[1]

wb = load_workbook(fp)
ws1 = wb['Sheet1']
ws2 = wb['blank']


# 編集するシートのデータを読み込み
values1 = [[cell.value for cell in row1] for row1 in ws1]

# 要らない行を削除
del values1[0]

#　フィルターしたい列を行に変更する
conv_value1 = list(zip(*values1))

# フィルターしたいデータに重複なデータを削除
customers_list = list(set(conv_value1[0]))
# 空白なデータを削除
customers_list.remove(None)

# データをloopして、同じお客さんのデータを集計する
for customers in customers_list:
	product_data = []
	for i in range(len(values1)):
		if values1[i][0] == customers:
			product_data.append(values1[i])
	ws3 = wb.create_sheet(customers)
	
	# データを各シートに書き込む
	for y,row in enumerate(product_data):
		for x,cell in enumerate(row):
			ws3.cell(row=y+2 , column=x+1 , value=product_data[y][x])

# Excelを保存する
wb.save(fp)

